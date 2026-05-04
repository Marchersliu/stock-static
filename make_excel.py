import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime, date

def get_weekday(dt):
    weekdays = ['周一','周二','周三','周四','周五','周六','周日']
    return weekdays[dt.weekday()]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "喷涂厂项目时间轴"

# ===== 样式定义 =====
# 表头样式
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 日期列
sub_header_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
sub_header_font = Font(name='微软雅黑', size=10, bold=True, color="000000")

# XDR 行样式
xdr_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 淡绿色
xdr_font = Font(name='微软雅黑', size=10, color="1B5E20")

# Darfon 行样式
darfon_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")  # 淡黄色
darfon_font = Font(name='微软雅黑', size=10, color="E65100")

# Holiday 行样式
holiday_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
holiday_font = Font(name='微软雅黑', size=10, italic=True, color="9E9E9E")

# 结果行样式
result_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 淡蓝色
result_font = Font(name='微软雅黑', size=10, color="0D47A1")

# 边框
thin_border = Border(
    left=Side(style='thin', color='BDBDBD'),
    right=Side(style='thin', color='BDBDBD'),
    top=Side(style='thin', color='BDBDBD'),
    bottom=Side(style='thin', color='BDBDBD')
)

# 通用对齐
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== 表头 =====
headers = ['日期', '星期', 'Substrate', '工序/任务', '描述/备注', '状态']
ws.append(headers)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# ===== 数据定义 =====
# 格式: (date_obj, weekday_cn, substrate, task, desc, status, style_type)
# style_type: 'xdr', 'darfon', 'holiday', 'result'

data = [
    # XDR
    (date(2026, 4, 29), '周二', 'XDR', 'Grey PR + Basecoat', 'XDR substrates 首次测试', '完成', 'xdr'),
    (date(2026, 4, 30), '周三', 'XDR', 'Laser + PU clear + UV clear', '激光处理 + 涂层固化', '完成', 'xdr'),
    (date(2026, 5, 1), '周四', '—', '劳动节假期', '工厂休假', '假期', 'holiday'),
    (date(2026, 5, 2), '周五', '—', '劳动节假期', '工厂休假', '假期', 'holiday'),
    (date(2026, 5, 3), '周六', '—', '劳动节假期', '工厂休假', '假期', 'holiday'),
    (date(2026, 5, 4), '周日', '—', '劳动节假期', '工厂休假', '假期', 'holiday'),
    (date(2026, 5, 5), '周一', '—', '劳动节假期', '工厂休假', '假期', 'holiday'),
    (date(2026, 5, 6), '周二', 'XDR', 'HWD results + THC 5cycles start', '高湿高热测试结果；THC 5cycles 启动', '进行中', 'result'),
    (date(2026, 5, 24), '周日', 'XDR', 'THC 5cycles results', 'THC 5cycles 测试完成，查看结果', '待完成', 'result'),

    # Darfon
    (date(2026, 5, 7), '周三', 'Darfon', 'Clear PR + Grey PR + Basecoat', '方案①：三层喷涂', '完成', 'darfon'),
    (date(2026, 5, 7), '周三', 'Darfon', 'Grey PR + Basecoat', '方案②：两层喷涂（同步进行）', '完成', 'darfon'),
    (date(2026, 5, 8), '周四', 'Darfon', 'Laser @Darfon', 'Darfon substrates 激光处理', '完成', 'darfon'),
    (date(2026, 5, 9), '周五', 'Darfon', 'UV clear', '方案①：UV clear 涂层', '完成', 'darfon'),
    (date(2026, 5, 9), '周五', 'Darfon', 'PU clear + UV clear', '方案②：PU clear + UV clear 涂层', '完成', 'darfon'),
    (date(2026, 5, 10), '周六', 'Darfon', 'UV clear', '方案①续：UV clear 固化', '完成', 'darfon'),
    (date(2026, 5, 10), '周六', 'Darfon', 'PU clear + UV clear', '方案②续：涂层固化', '完成', 'darfon'),
    (date(2026, 5, 11), '周日', 'Darfon', 'THC 5cycles start', 'Darfon THC 5cycles 启动', '进行中', 'result'),
    (date(2026, 5, 12), '周一', 'Darfon', 'HWD results', '高湿高热测试结果', '待完成', 'result'),
    (date(2026, 5, 29), '周四', 'Darfon', 'THC 5cycles results', 'THC 5cycles 测试完成，查看结果', '待完成', 'result'),
]

# ===== 填充数据 =====
row_num = 2
for dt, wd, substrate, task, desc, status, style_type in data:
    ws.append([
        dt.strftime('%Y-%m-%d'),
        wd,
        substrate,
        task,
        desc,
        status
    ])
    
    # 应用样式
    if style_type == 'xdr':
        fill = xdr_fill
        font = xdr_font
    elif style_type == 'darfon':
        fill = darfon_fill
        font = darfon_font
    elif style_type == 'holiday':
        fill = holiday_fill
        font = holiday_font
    elif style_type == 'result':
        fill = result_fill
        font = result_font
    else:
        fill = None
        font = Font(name='微软雅黑', size=10)
    
    for col in range(1, 7):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = left_align if col in [4, 5] else center_align
    
    # 日期列居中
    ws.cell(row=row_num, column=1).alignment = center_align
    
    row_num += 1

# ===== 合并假日单元格（视觉更清晰） =====
# 合并 5/1~5/5 的日期行
ws.merge_cells(start_row=4, start_column=1, end_row=8, end_column=1)
ws.cell(row=4, column=1).value = "2026-05-01\n~\n2026-05-05"
ws.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 合并5/7的两个Darfon任务（Substrate列）
ws.merge_cells(start_row=12, start_column=3, end_row=13, end_column=3)
ws.cell(row=12, column=3).value = "Darfon"
ws.cell(row=12, column=3).alignment = center_align

# 合并5/9的两个任务
ws.merge_cells(start_row=14, start_column=1, end_row=15, end_column=1)
ws.cell(row=14, column=1).value = "2026-05-09\n（周五）"
ws.cell(row=14, column=1).alignment = center_align

# 合并5/10的两个任务
ws.merge_cells(start_row=16, start_column=1, end_row=17, end_column=1)
ws.cell(row=16, column=1).value = "2026-05-10\n（周六）"
ws.cell(row=16, column=1).alignment = center_align

# ===== 列宽设置 =====
column_widths = [14, 8, 12, 28, 40, 10]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ===== 行高 =====
ws.row_dimensions[1].height = 28
for r in range(2, row_num):
    ws.row_dimensions[r].height = 32

# ===== 冻结首行 =====
ws.freeze_panes = 'A2'

# ===== 添加汇总Sheet =====
ws2 = wb.create_sheet("项目汇总")

ws2.append(['Substrate', '关键里程碑', '日期', '说明'])
for col, header in enumerate(['Substrate', '关键里程碑', '日期', '说明'], 1):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

milestones = [
    ('XDR', 'Grey PR + Basecoat 测试', '2026-04-29', '首次喷涂测试'),
    ('XDR', 'Laser + PU clear + UV clear', '2026-04-30', '激光+涂层处理'),
    ('XDR', 'HWD results + THC start', '2026-05-06', '高湿测试+THC启动'),
    ('XDR', 'THC 5cycles results', '2026-05-24', '最终测试结果'),
    ('Darfon', 'Clear PR + Grey PR + Basecoat', '2026-05-07', '方案①三层喷涂'),
    ('Darfon', 'Grey PR + Basecoat', '2026-05-07', '方案②两层喷涂'),
    ('Darfon', 'Laser @Darfon', '2026-05-08', '激光处理'),
    ('Darfon', 'UV clear / PU+UV clear', '2026-05-09~10', '涂层固化'),
    ('Darfon', 'THC 5cycles start', '2026-05-11', 'THC启动'),
    ('Darfon', 'HWD results', '2026-05-12', '高湿测试结果'),
    ('Darfon', 'THC 5cycles results', '2026-05-29', '最终测试结果'),
]

for i, (sub, milestone, dt, note) in enumerate(milestones, 2):
    ws2.append([sub, milestone, dt, note])
    for col in range(1, 5):
        cell = ws2.cell(row=i, column=col)
        cell.border = thin_border
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = left_align if col in [2, 4] else center_align
        if sub == 'XDR':
            cell.fill = xdr_fill
        elif sub == 'Darfon':
            cell.fill = darfon_fill

for i, width in enumerate([12, 32, 16, 40], 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

ws2.row_dimensions[1].height = 28
ws2.freeze_panes = 'A2'

# ===== 保存 =====
output_path = '/Users/hf/.kimi_openclaw/workspace/喷涂厂项目时间轴.xlsx'
wb.save(output_path)
print(f"✅ Excel 已创建: {output_path}")
