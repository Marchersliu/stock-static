import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "喷涂厂时间轴"

# ===== 样式 =====
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name='微软雅黑', size=10, bold=True, color="FFFFFF")
sub_header_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
sub_header_font = Font(name='微软雅黑', size=9, bold=True, color="000000")

xdr_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
xdr_font = Font(name='微软雅黑', size=9, color="1B5E20")

darfon_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
darfon_font = Font(name='微软雅黑', size=9, color="E65100")

holiday_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
holiday_font = Font(name='微软雅黑', size=9, italic=True, color="9E9E9E")

result_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
result_font = Font(name='微软雅黑', size=9, color="0D47A1", bold=True)

today_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='BDBDBD'),
    right=Side(style='thin', color='BDBDBD'),
    top=Side(style='thin', color='BDBDBD'),
    bottom=Side(style='thin', color='BDBDBD')
)

# ===== 数据准备 =====
# 日期范围: 4/29 ~ 5/29
start = date(2026, 4, 29)
end = date(2026, 5, 29)
dates = []
d = start
while d <= end:
    dates.append(d)
    d += timedelta(days=1)

weekdays = ['周一','周二','周三','周四','周五','周六','周日']

# 任务映射: {date_obj: [task1, task2, ...]}
xdr_tasks = {}
darfon_tasks = {}

# XDR
xdr_tasks[date(2026, 4, 29)] = ['Grey PR\n+ Basecoat']
xdr_tasks[date(2026, 4, 30)] = ['Laser\n+ PU clear\n+ UV clear']
xdr_tasks[date(2026, 5, 1)] = ['🏖️ 劳动节\n假期']
xdr_tasks[date(2026, 5, 2)] = ['🏖️ 劳动节\n假期']
xdr_tasks[date(2026, 5, 3)] = ['🏖️ 劳动节\n假期']
xdr_tasks[date(2026, 5, 4)] = ['🏖️ 劳动节\n假期']
xdr_tasks[date(2026, 5, 5)] = ['🏖️ 劳动节\n假期']
xdr_tasks[date(2026, 5, 6)] = ['HWD results', 'THC 5cycles\nstart']
xdr_tasks[date(2026, 5, 24)] = ['THC 5cycles\nresults 🎯']

# Darfon
darfon_tasks[date(2026, 5, 1)] = ['🏖️ 劳动节\n假期']
darfon_tasks[date(2026, 5, 2)] = ['🏖️ 劳动节\n假期']
darfon_tasks[date(2026, 5, 3)] = ['🏖️ 劳动节\n假期']
darfon_tasks[date(2026, 5, 4)] = ['🏖️ 劳动节\n假期']
darfon_tasks[date(2026, 5, 5)] = ['🏖️ 劳动节\n假期']
darfon_tasks[date(2026, 5, 7)] = ['① Clear PR\n+ Grey PR\n+ Basecoat', '② Grey PR\n+ Basecoat']
darfon_tasks[date(2026, 5, 8)] = ['Laser\n@Darfon']
darfon_tasks[date(2026, 5, 9)] = ['① UV clear', '② PU clear\n+ UV clear']
darfon_tasks[date(2026, 5, 10)] = ['① UV clear', '② PU clear\n+ UV clear']
darfon_tasks[date(2026, 5, 11)] = ['THC 5cycles\nstart']
darfon_tasks[date(2026, 5, 12)] = ['HWD results']
darfon_tasks[date(2026, 5, 29)] = ['THC 5cycles\nresults 🎯']

# ===== 写入表头 =====
# Row 1: Substrate label + dates
ws.cell(row=1, column=1, value="Substrate")
ws.cell(row=1, column=1).fill = header_fill
ws.cell(row=1, column=1).font = header_font
ws.cell(row=1, column=1).alignment = center_align
ws.cell(row=1, column=1).border = thin_border

for i, dt in enumerate(dates, 2):
    cell = ws.cell(row=1, column=i, value=dt.strftime('%m/%d'))
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    # 周末列加深一点
    if dt.weekday() >= 5:
        cell.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")

# Row 2: 星期
ws.cell(row=2, column=1, value="星期")
ws.cell(row=2, column=1).fill = sub_header_fill
ws.cell(row=2, column=1).font = sub_header_font
ws.cell(row=2, column=1).alignment = center_align
ws.cell(row=2, column=1).border = thin_border

for i, dt in enumerate(dates, 2):
    wd = weekdays[dt.weekday()]
    cell = ws.cell(row=2, column=i, value=wd)
    cell.fill = sub_header_fill
    cell.font = sub_header_font
    cell.alignment = center_align
    cell.border = thin_border
    if dt.weekday() >= 5:
        cell.fill = PatternFill(start_color="C5D1E0", end_color="C5D1E0", fill_type="solid")

# Row 3: XDR
ws.cell(row=3, column=1, value="XDR\nsubstrates")
ws.cell(row=3, column=1).fill = xdr_fill
ws.cell(row=3, column=1).font = Font(name='微软雅黑', size=10, bold=True, color="1B5E20")
ws.cell(row=3, column=1).alignment = center_align
ws.cell(row=3, column=1).border = thin_border

for i, dt in enumerate(dates, 2):
    tasks = xdr_tasks.get(dt, [])
    text = '\n'.join(tasks) if tasks else ''
    cell = ws.cell(row=3, column=i, value=text)
    cell.border = thin_border
    cell.alignment = center_align
    
    if dt in xdr_tasks:
        if '🏖️' in text:
            cell.fill = holiday_fill
            cell.font = holiday_font
        elif 'results' in text or 'THC' in text:
            cell.fill = result_fill
            cell.font = result_font
        else:
            cell.fill = xdr_fill
            cell.font = xdr_font
    else:
        cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

# Row 4: Darfon
ws.cell(row=4, column=1, value="Darfon\nsubstrates")
ws.cell(row=4, column=1).fill = darfon_fill
ws.cell(row=4, column=1).font = Font(name='微软雅黑', size=10, bold=True, color="E65100")
ws.cell(row=4, column=1).alignment = center_align
ws.cell(row=4, column=1).border = thin_border

for i, dt in enumerate(dates, 2):
    tasks = darfon_tasks.get(dt, [])
    text = '\n'.join(tasks) if tasks else ''
    cell = ws.cell(row=4, column=i, value=text)
    cell.border = thin_border
    cell.alignment = center_align
    
    if dt in darfon_tasks:
        if '🏖️' in text:
            cell.fill = holiday_fill
            cell.font = holiday_font
        elif 'results' in text or 'THC' in text:
            cell.fill = result_fill
            cell.font = result_font
        else:
            cell.fill = darfon_fill
            cell.font = darfon_font
    else:
        cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

# ===== 列宽 =====
ws.column_dimensions['A'].width = 14
for i in range(2, len(dates)+2):
    ws.column_dimensions[get_column_letter(i)].width = 11

# ===== 行高 =====
ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 65
ws.row_dimensions[4].height = 65

# ===== 合并假日列（4/29~5/5 假期背景标注）=====
# 在顶部加一行小标题说明
ws.insert_rows(1)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
ws.cell(row=1, column=1, value="喷涂厂项目时间轴（4/29 ~ 5/29）")
ws.cell(row=1, column=1).font = Font(name='微软雅黑', size=14, bold=True, color="1F4E79")
ws.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(dates)+1)
ws.cell(row=1, column=2, value="← 时间从左到右 →   🟢 XDR   🟡 Darfon   🔵 测试结果   🏖️ 假期")
ws.cell(row=1, column=2).font = Font(name='微软雅黑', size=10, color="666666")
ws.cell(row=1, column=2).alignment = Alignment(horizontal='right', vertical='center')

ws.row_dimensions[1].height = 30

# ===== Sheet2: 关键里程碑 =====
ws2 = wb.create_sheet("关键里程碑")

headers2 = ['日期', '星期', 'Substrate', '里程碑', '说明']
ws2.append(headers2)
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

milestones = [
    (date(2026, 4, 29), '周二', 'XDR', 'Grey PR + Basecoat 测试', '首次喷涂'),
    (date(2026, 4, 30), '周三', 'XDR', 'Laser + PU clear + UV clear', '激光+涂层'),
    (date(2026, 5, 6), '周二', 'XDR', 'HWD results + THC start', '高湿测试+THC启动'),
    (date(2026, 5, 24), '周日', 'XDR', 'THC 5cycles results', '最终测试 🎯'),
    (date(2026, 5, 7), '周三', 'Darfon', '方案①: Clear+Grey PR+Basecoat', '三层喷涂'),
    (date(2026, 5, 7), '周三', 'Darfon', '方案②: Grey PR+Basecoat', '两层喷涂'),
    (date(2026, 5, 8), '周四', 'Darfon', 'Laser @Darfon', '激光处理'),
    (date(2026, 5, 9), '周五', 'Darfon', '方案①: UV clear', '涂层固化'),
    (date(2026, 5, 9), '周五', 'Darfon', '方案②: PU clear + UV clear', '涂层固化'),
    (date(2026, 5, 11), '周日', 'Darfon', 'THC 5cycles start', 'THC启动'),
    (date(2026, 5, 12), '周一', 'Darfon', 'HWD results', '高湿测试'),
    (date(2026, 5, 29), '周四', 'Darfon', 'THC 5cycles results', '最终测试 🎯'),
]

for i, (dt, wd, sub, ms, note) in enumerate(milestones, 2):
    ws2.append([dt.strftime('%Y-%m-%d'), wd, sub, ms, note])
    for col in range(1, 6):
        cell = ws2.cell(row=i, column=col)
        cell.border = thin_border
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = left_align if col in [4, 5] else center_align
        if sub == 'XDR':
            cell.fill = xdr_fill
        elif sub == 'Darfon':
            cell.fill = darfon_fill

for i, w in enumerate([12, 8, 12, 32, 28], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 28
ws2.freeze_panes = 'A2'

# ===== 保存 =====
output = '/Users/hf/.kimi_openclaw/workspace/喷涂厂项目时间轴.xlsx'
wb.save(output)
print(f"✅ Excel 已创建: {output}")
