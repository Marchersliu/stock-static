import openpyxl
from PIL import Image, ImageDraw, ImageFont
import textwrap

# ===== 读取Excel数据 =====
wb = openpyxl.load_workbook('/Users/hf/.kimi_openclaw/workspace/喷涂厂项目时间轴.xlsx')
ws = wb['喷涂厂时间轴']

# 读取时间轴数据
rows = []
for r in range(1, 6):
    row_data = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=r, column=c).value
        row_data.append(val if val is not None else "")
    rows.append(row_data)

# ===== 图片参数 =====
CELL_W = 100
CELL_H = 40
HEADER_H = 50
COL_A_W = 130
ROW_COUNT = len(rows)
COL_COUNT = len(rows[0])

IMG_W = COL_A_W + (COL_COUNT - 1) * CELL_W + 80
IMG_H = HEADER_H + ROW_COUNT * CELL_H + 300  # extra for summary text

# 颜色
COLORS = {
    'title': '#1F4E79',
    'header': '#1F4E79',
    'sub': '#D6DCE4',
    'xdr': '#E8F5E9',
    'xdr_header': '#C8E6C9',
    'darfon': '#FFF8E1',
    'darfon_header': '#FFECB3',
    'holiday': '#ECEFF1',
    'blank': '#FAFAFA',
    'result': '#E3F2FD',
    'text_dark': '#333333',
    'text_white': '#FFFFFF',
    'text_gray': '#78909C',
    'border': '#BDBDBD',
}

img = Image.new('RGB', (IMG_W, IMG_H), '#FFFFFF')
draw = ImageDraw.Draw(img)

# 字体
try:
    font_title = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 22)
    font_header = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 13)
    font_cell = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 11)
    font_small = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 10)
    font_summary = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 12)
except:
    font_title = ImageFont.load_default()
    font_header = ImageFont.load_default()
    font_cell = ImageFont.load_default()
    font_small = ImageFont.load_default()
    font_summary = ImageFont.load_default()

# ===== 绘制标题 =====
title_y = 20
draw.text((40, title_y), "喷涂厂项目时间轴（XDR & Darfon Substrates）", fill=COLORS['title'], font=font_title)

# ===== 绘制表格 =====
table_y = title_y + 45

# 计算每列x位置
col_x = [40]
for i in range(1, COL_COUNT):
    w = COL_A_W if i == 0 else CELL_W
    col_x.append(col_x[-1] + w)

# 行高
row_heights = [40, 35, 35, 65, 65]

# 绘制单元格
for r, row_data in enumerate(rows):
    y = table_y + sum(row_heights[:r])
    h = row_heights[r]
    
    for c, val in enumerate(row_data):
        x = col_x[c]
        w = COL_A_W if c == 0 else CELL_W
        
        # 判断颜色
        fill_color = '#FFFFFF'
        text_color = COLORS['text_dark']
        
        if r == 0:
            fill_color = COLORS['title']
            text_color = COLORS['text_white']
        elif r == 1:
            if c == 0:
                fill_color = COLORS['header']
                text_color = COLORS['text_white']
            elif val == '05/01\n~\n05/05' or val == '...':
                fill_color = COLORS['holiday']
                text_color = COLORS['text_gray']
            else:
                fill_color = COLORS['header']
                text_color = COLORS['text_white']
        elif r == 2:
            if c == 0:
                fill_color = COLORS['sub']
                text_color = COLORS['text_dark']
            elif val == 'Holiday' or val == '...':
                fill_color = COLORS['holiday']
                text_color = COLORS['text_gray']
            else:
                fill_color = COLORS['sub']
                text_color = COLORS['text_dark']
        elif r == 3:
            if c == 0:
                fill_color = COLORS['xdr_header']
                text_color = '#1B5E20'
            elif val == 'Holiday':
                fill_color = COLORS['holiday']
                text_color = COLORS['text_gray']
            elif val == '':
                fill_color = COLORS['blank']
                text_color = COLORS['text_gray']
            elif 'results' in str(val):
                fill_color = COLORS['result']
                text_color = '#0D47A1'
            else:
                fill_color = COLORS['xdr']
                text_color = '#1B5E20'
        elif r == 4:
            if c == 0:
                fill_color = COLORS['darfon_header']
                text_color = '#E65100'
            elif val == 'Holiday':
                fill_color = COLORS['holiday']
                text_color = COLORS['text_gray']
            elif val == '':
                fill_color = COLORS['blank']
                text_color = COLORS['text_gray']
            elif 'results' in str(val):
                fill_color = COLORS['result']
                text_color = '#0D47A1'
            else:
                fill_color = COLORS['darfon']
                text_color = '#E65100'
        
        # 绘制背景
        draw.rectangle([x, y, x + w, y + h], fill=fill_color, outline=COLORS['border'], width=1)
        
        # 绘制文字
        if val:
            text = str(val)
            # 计算文字位置（居中）
            bbox = draw.textbbox((0, 0), text, font=font_cell if r >= 3 else font_header)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            tx = x + (w - tw) // 2
            ty = y + (h - th) // 2
            
            # 多行文字处理
            if '\n' in text:
                lines = text.split('\n')
                line_h = th // len(lines) if lines else 14
                start_y = y + (h - line_h * len(lines)) // 2
                for i_line, line in enumerate(lines):
                    lb = draw.textbbox((0, 0), line, font=font_cell)
                    lw = lb[2] - lb[0]
                    lx = x + (w - lw) // 2
                    ly = start_y + i_line * line_h
                    draw.text((lx, ly), line, fill=text_color, font=font_cell)
            else:
                draw.text((tx, ty), text, fill=text_color, font=font_cell if r >= 3 else font_header)

# ===== 绘制总结区域 =====
summary_y = table_y + sum(row_heights) + 30

summary_text = [
    "📋 协作总结",
    "",
    "用户：把喷涂厂两个substrates（XDR + Darfon）的项目按时间轴做成Excel，",
    "要求：XDR和Darfon上下并排，时间从左到右，假期和空白段合并成等宽单元格。",
    "",
    "我：用Python + openpyxl生成了时间轴Excel，反复调整3版：",
    "  → 第1版：31天逐列，太宽",
    "  → 第2版：假期合并但宽度不一致",
    "  → 第3版：假期+空白段都合并，所有列等宽，一屏看完",
    "",
    "用户：可以了，你帮忙把咱们的对话，总结一下，你的回复太多内容，挑简要的，",
    "然后再把你做的excel文件截屏，总结成一张照片，我分享给我同事看看，",
    "如何和你一起工作。把我这句话也放入总结中。总结成pdf也行。",
    "",
    "我：生成这张总结图片（包含Excel截图+协作过程简述）。",
    "",
    "💡 协作特点：自然语言描述需求 → 快速迭代 → 精确交付"
]

x_text = 40
y_text = summary_y
for line in summary_text:
    draw.text((x_text, y_text), line, fill='#333333', font=font_summary)
    y_text += 22

# 保存
output = '/Users/hf/.kimi_openclaw/workspace/协作总结_喷涂厂时间轴.png'
img.save(output, quality=95)
print(f"✅ 总结图片已保存: {output}")

# 也生成PDF版本
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

try:
    pdfmetrics.registerFont(TTFont('PingFang', '/System/Library/Fonts/PingFang.ttc'))
    font_name = 'PingFang'
except:
    font_name = 'Helvetica'

pdf_path = '/Users/hf/.kimi_openclaw/workspace/协作总结_喷涂厂时间轴.pdf'
c = pdf_canvas.Canvas(pdf_path, pagesize=A4)
width, height = A4

# 标题
c.setFont(font_name, 16)
c.drawString(20*mm, height - 20*mm, "喷涂厂项目时间轴（XDR & Darfon Substrates）")

# 插入图片
c.drawImage(output, 20*mm, height - 180*mm, width=170*mm, preserveAspectRatio=True, anchor='n')

# 总结文字
c.setFont(font_name, 10)
y = height - 200*mm
for line in summary_text:
    c.drawString(20*mm, y, line)
    y -= 6*mm
    if y < 20*mm:
        c.showPage()
        y = height - 20*mm

c.save()
print(f"✅ 总结PDF已保存: {pdf_path}")

# 同步到iCloud
import shutil
shutil.copy(output, '/Users/hf/Library/Mobile Documents/com~apple~CloudDocs/下载文件/HF/协作总结_喷涂厂时间轴.png')
shutil.copy(pdf_path, '/Users/hf/Library/Mobile Documents/com~apple~CloudDocs/下载文件/HF/协作总结_喷涂厂时间轴.pdf')
print("✅ 已同步到 iCloud")
