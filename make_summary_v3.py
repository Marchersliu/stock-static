from PIL import Image, ImageDraw, ImageFont
import openpyxl

# ===== 读取Excel数据 =====
wb = openpyxl.load_workbook('/Users/hf/.kimi_openclaw/workspace/喷涂厂项目时间轴.xlsx')
ws = wb['喷涂厂时间轴']

rows = []
for r in range(1, 6):
    row_data = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=r, column=c).value
        row_data.append(val if val is not None else "")
    rows.append(row_data)

# ===== 字体（用Hiragino，PingFang加载有问题）=====
FONT_PATH = '/System/Library/Fonts/Hiragino Sans GB.ttc'
font_title = ImageFont.truetype(FONT_PATH, 26)
font_header = ImageFont.truetype(FONT_PATH, 14)
font_cell = ImageFont.truetype(FONT_PATH, 13)
font_summary = ImageFont.truetype(FONT_PATH, 16)
font_bold = ImageFont.truetype(FONT_PATH, 16)

# ===== 图片参数 =====
CELL_W = 115
COL_A_W = 150
MARGIN = 70

COL_COUNT = len(rows[0])
ROW_COUNT = len(rows)

TABLE_W = COL_A_W + (COL_COUNT - 1) * CELL_W
IMG_W = TABLE_W + MARGIN * 2

# 表格行高
row_heights = [50, 42, 42, 80, 80]
TABLE_H = sum(row_heights)

# 总结区域估算高度
summary_line_h = 28
summary_lines = [
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
    "",
    "📋 协作总结",
    "",
    "用户：把喷涂厂两个substrates（XDR + Darfon）的项目按时间轴做成Excel，",
    "要求：XDR和Darfon上下并排，时间从左到右，假期和空白段合并成等宽单元格。",
    "",
    "我：用Python + openpyxl生成时间轴Excel，反复调整3版：",
    "    第1版：31天逐列，太宽",
    "    第2版：假期合并但宽度不一致", 
    "    第3版：假期+空白段都合并，所有列等宽，一屏看完",
    "",
    "用户：可以了，你帮忙把咱们的对话，总结一下，你的回复太多内容，挑简要的，",
    "然后再把你做的excel文件截屏，总结成一张照片，我分享给我同事看看，",
    "如何和你一起工作。把我这句话也放入总结中。总结成pdf也行。",
    "",
    "我：生成这张总结图片（包含Excel截图+协作过程简述）。",
    "",
    "💡 协作特点：自然语言描述需求 → 快速迭代 → 精确交付",
    "",
    "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
]

SUMMARY_H = len(summary_lines) * summary_line_h + 60
IMG_H = 40 + TABLE_H + SUMMARY_H + 40

# ===== 创建画布 =====
img = Image.new('RGB', (IMG_W, IMG_H), '#FFFFFF')
draw = ImageDraw.Draw(img)

# ===== 颜色 =====
C = {
    'title': '#1F4E79',
    'header': '#1F4E79',
    'sub': '#D6DCE4',
    'xdr': '#E8F5E9',
    'xdr_h': '#C8E6C9',
    'darfon': '#FFF8E1',
    'darfon_h': '#FFECB3',
    'holiday': '#ECEFF1',
    'blank': '#FAFAFA',
    'result': '#E3F2FD',
    'dark': '#333333',
    'white': '#FFFFFF',
    'gray': '#78909C',
    'border': '#BDBDBD',
}

# ===== 标题 =====
draw.text((MARGIN, 25), "喷涂厂项目时间轴（XDR & Darfon Substrates）", fill=C['title'], font=font_title)

# ===== 表格 =====
table_y = 80
col_x = [MARGIN]
for i in range(1, COL_COUNT):
    col_x.append(col_x[-1] + (COL_A_W if i == 0 else CELL_W))

for r, row_data in enumerate(rows):
    y = table_y + sum(row_heights[:r])
    h = row_heights[r]
    
    for c, val in enumerate(row_data):
        x = col_x[c]
        w = COL_A_W if c == 0 else CELL_W
        
        # 颜色判断
        fill_color = '#FFFFFF'
        text_color = C['dark']
        
        if r == 0:
            fill_color = C['title']
            text_color = C['white']
        elif r == 1:
            if c == 0:
                fill_color = C['header']
                text_color = C['white']
            elif val in ('05/01\n~\n05/05', '...'):
                fill_color = C['holiday']
                text_color = C['gray']
            else:
                fill_color = C['header']
                text_color = C['white']
        elif r == 2:
            if c == 0:
                fill_color = C['sub']
                text_color = C['dark']
            elif val in ('Holiday', '...'):
                fill_color = C['holiday']
                text_color = C['gray']
            else:
                fill_color = C['sub']
                text_color = C['dark']
        elif r == 3:
            if c == 0:
                fill_color = C['xdr_h']
                text_color = '#1B5E20'
            elif val == 'Holiday':
                fill_color = C['holiday']
                text_color = C['gray']
            elif val == '':
                fill_color = C['blank']
                text_color = C['gray']
            elif 'results' in str(val):
                fill_color = C['result']
                text_color = '#0D47A1'
            else:
                fill_color = C['xdr']
                text_color = '#1B5E20'
        elif r == 4:
            if c == 0:
                fill_color = C['darfon_h']
                text_color = '#E65100'
            elif val == 'Holiday':
                fill_color = C['holiday']
                text_color = C['gray']
            elif val == '':
                fill_color = C['blank']
                text_color = C['gray']
            elif 'results' in str(val):
                fill_color = C['result']
                text_color = '#0D47A1'
            else:
                fill_color = C['darfon']
                text_color = '#E65100'
        
        draw.rectangle([x, y, x + w, y + h], fill=fill_color, outline=C['border'], width=1)
        
        if val:
            text = str(val)
            if '\n' in text:
                lines = text.split('\n')
                bbox = draw.textbbox((0, 0), 'Ay', font=font_cell)
                line_h = bbox[3] - bbox[1] + 4
                total_h = line_h * len(lines)
                start_y = y + (h - total_h) // 2
                for i_line, line in enumerate(lines):
                    lb = draw.textbbox((0, 0), line, font=font_cell)
                    lw = lb[2] - lb[0]
                    lx = x + (w - lw) // 2
                    ly = start_y + i_line * line_h
                    draw.text((lx, ly), line, fill=text_color, font=font_cell)
            else:
                bbox = draw.textbbox((0, 0), text, font=font_cell if r >= 3 else font_header)
                tw = bbox[2] - bbox[0]
                th = bbox[3] - bbox[1]
                tx = x + (w - tw) // 2
                ty = y + (h - th) // 2
                draw.text((tx, ty), text, fill=text_color, font=font_cell if r >= 3 else font_header)

# ===== 总结文字 =====
summary_y = table_y + TABLE_H + 50
y_text = summary_y
for line in summary_lines:
    draw.text((MARGIN, y_text), line, fill='#333333', font=font_summary)
    y_text += summary_line_h

# 裁剪
final_h = y_text + 40
img = img.crop((0, 0, IMG_W, final_h))

# ===== 保存 =====
png_path = '/Users/hf/.kimi_openclaw/workspace/协作总结_喷涂厂时间轴.png'
img.save(png_path, quality=95)
print(f"✅ PNG: {png_path}")

pdf_path = '/Users/hf/.kimi_openclaw/workspace/协作总结_喷涂厂时间轴.pdf'
img.save(pdf_path, "PDF", resolution=150.0)
print(f"✅ PDF: {pdf_path}")

import shutil
shutil.copy(png_path, '/Users/hf/Library/Mobile Documents/com~apple~CloudDocs/下载文件/HF/协作总结_喷涂厂时间轴.png')
shutil.copy(pdf_path, '/Users/hf/Library/Mobile Documents/com~apple~CloudDocs/下载文件/HF/协作总结_喷涂厂时间轴.pdf')
print("✅ iCloud同步完成")
